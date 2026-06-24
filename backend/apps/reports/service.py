"""
ReportService - Rapports analytiques multi-tenant.
- Dashboard temps reel
- Resume par periode
- Export CSV / Excel / PDF
- Business intelligence client et fidelite
"""

import csv
import io
from calendar import monthrange
from datetime import date, datetime, timedelta, timezone
from typing import Optional
from uuid import UUID

from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from apps.loyalty.models import CustomerScore, LoyaltyCoupon, LoyaltyTransaction
from apps.orders.models import Order, OrderItem
from apps.payments.models import Payment
from apps.users.models import User


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    def _period_bounds(self, period: str) -> tuple[datetime, datetime]:
        now = datetime.now(timezone.utc)
        if period == "today":
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == "week":
            start = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        elif period == "month":
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        elif period == "year":
            start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return start, now

    async def _get_segment_counts(self, company_id: UUID) -> dict[str, int]:
        result = await self.db.execute(
            select(CustomerScore.segment, func.count(CustomerScore.id))
            .where(CustomerScore.company_id == company_id)
            .group_by(CustomerScore.segment)
        )
        counts = {segment: 0 for segment in ("new", "active", "loyal", "vip", "at_risk", "inactive")}
        for segment, count in result.all():
            counts[segment] = int(count or 0)
        return counts

    async def _get_top_customers(
        self,
        company_id: UUID,
        start: datetime,
        end: datetime,
        limit: int = 5,
    ) -> list[dict]:
        result = await self.db.execute(
            select(
                Payment.customer_id,
                User.first_name,
                User.last_name,
                func.count(Payment.id).label("orders_count"),
                func.coalesce(func.sum(Payment.amount_xof), 0).label("total_spent_xof"),
                CustomerScore.segment,
            )
            .join(Order, Order.id == Payment.order_id)
            .join(User, User.id == Payment.customer_id)
            .outerjoin(
                CustomerScore,
                and_(
                    CustomerScore.company_id == Order.company_id,
                    CustomerScore.customer_id == Payment.customer_id,
                ),
            )
            .where(
                Order.company_id == company_id,
                Payment.status == "confirmed",
                Payment.confirmed_at.between(start, end),
            )
            .group_by(
                Payment.customer_id,
                User.first_name,
                User.last_name,
                CustomerScore.segment,
            )
            .order_by(func.coalesce(func.sum(Payment.amount_xof), 0).desc())
            .limit(limit)
        )
        return [
            {
                "customer_id": str(row.customer_id),
                "customer_name": f"{row.first_name} {row.last_name}".strip(),
                "orders_count": int(row.orders_count or 0),
                "total_spent_xof": int(row.total_spent_xof or 0),
                "segment": row.segment or "new",
            }
            for row in result.all()
        ]

    async def get_dashboard(self, company_id: UUID) -> dict:
        now = datetime.now(timezone.utc)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        rolling_start = now - timedelta(days=30)

        orders_today = await self.db.scalar(
            select(func.count(Order.id)).where(
                Order.company_id == company_id,
                Order.created_at >= today_start,
            )
        ) or 0

        pending_orders = await self.db.scalar(
            select(func.count(Order.id)).where(
                Order.company_id == company_id,
                Order.status.in_(["confirmed", "preparing"]),
            )
        ) or 0

        pending_payments = await self.db.scalar(
            select(func.count(Payment.id))
            .join(Order, Order.id == Payment.order_id)
            .where(Order.company_id == company_id, Payment.status == "pending_verification")
        ) or 0

        revenue_today = await self.db.scalar(
            select(func.sum(Payment.amount_xof))
            .join(Order, Order.id == Payment.order_id)
            .where(
                Order.company_id == company_id,
                Payment.status == "confirmed",
                Payment.confirmed_at >= today_start,
            )
        ) or 0

        active_customers_30d = await self.db.scalar(
            select(func.count(func.distinct(Order.customer_id))).where(
                Order.company_id == company_id,
                Order.created_at >= rolling_start,
            )
        ) or 0

        return {
            "orders_today": int(orders_today),
            "pending_orders": int(pending_orders),
            "pending_payments": int(pending_payments),
            "revenue_today_xof": int(revenue_today),
            "avg_order_today_xof": round(revenue_today / orders_today) if orders_today else 0,
            "active_customers_30d": int(active_customers_30d),
            "customer_segments": await self._get_segment_counts(company_id),
            "top_customers": await self._get_top_customers(company_id, rolling_start, now),
        }

    async def get_summary(
        self,
        company_id: UUID,
        period: str = "today",
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
    ) -> dict:
        if date_from and date_to:
            start = datetime.combine(date_from, datetime.min.time()).replace(tzinfo=timezone.utc)
            end = datetime.combine(date_to, datetime.max.time()).replace(tzinfo=timezone.utc)
        else:
            start, end = self._period_bounds(period)

        base_filter = and_(
            Order.company_id == company_id,
            Order.created_at >= start,
            Order.created_at <= end,
        )

        orders_count = await self.db.scalar(select(func.count(Order.id)).where(base_filter)) or 0
        orders_delivered = await self.db.scalar(
            select(func.count(Order.id)).where(base_filter, Order.status == "delivered")
        ) or 0
        orders_cancelled = await self.db.scalar(
            select(func.count(Order.id)).where(base_filter, Order.status == "cancelled")
        ) or 0

        revenue_result = await self.db.scalar(
            select(func.sum(Payment.amount_xof))
            .join(Order, Order.id == Payment.order_id)
            .where(
                Order.company_id == company_id,
                Payment.status == "confirmed",
                Payment.confirmed_at.between(start, end),
            )
        ) or 0

        avg_order = (revenue_result / orders_delivered) if orders_delivered > 0 else 0

        unique_customers = await self.db.scalar(
            select(func.count(func.distinct(Order.customer_id))).where(base_filter)
        ) or 0

        top_products_result = await self.db.execute(
            select(
                OrderItem.product_id,
                OrderItem.product_name,
                func.sum(OrderItem.quantity).label("qty"),
                func.sum(OrderItem.subtotal_xof).label("revenue_xof"),
            )
            .join(Order, Order.id == OrderItem.order_id)
            .where(base_filter, Order.status == "delivered")
            .group_by(OrderItem.product_id, OrderItem.product_name)
            .order_by(func.sum(OrderItem.quantity).desc())
            .limit(10)
        )
        top_products = [
            {
                "product_id": str(row.product_id) if row.product_id else None,
                "product_name": row.product_name,
                "name": row.product_name,
                "quantity_sold": int(row.qty or 0),
                "revenue_xof": int(row.revenue_xof or 0),
            }
            for row in top_products_result.all()
        ]

        payment_ops = await self.db.execute(
            select(
                Payment.operator,
                func.count(Payment.id).label("count"),
                func.sum(Payment.amount_xof).label("total"),
            )
            .join(Order, Order.id == Payment.order_id)
            .where(
                Order.company_id == company_id,
                Payment.status == "confirmed",
                Payment.confirmed_at.between(start, end),
            )
            .group_by(Payment.operator)
        )
        payment_by_operator = [
            {"operator": row.operator, "count": int(row.count or 0), "total_xof": int(row.total or 0)}
            for row in payment_ops.all()
        ]

        new_customers = await self.db.scalar(
            select(func.count())
            .select_from(
                select(Order.customer_id)
                .where(base_filter)
                .group_by(Order.customer_id)
                .having(func.min(Order.created_at).between(start, end))
                .subquery()
            )
        ) or 0

        points_distributed = await self.db.scalar(
            select(func.coalesce(func.sum(LoyaltyTransaction.points_delta), 0)).where(
                LoyaltyTransaction.company_id == company_id,
                LoyaltyTransaction.created_at.between(start, end),
                LoyaltyTransaction.points_delta > 0,
            )
        ) or 0

        coupons_used = await self.db.scalar(
            select(func.count(LoyaltyCoupon.id)).where(
                LoyaltyCoupon.company_id == company_id,
                LoyaltyCoupon.is_used == True,
                LoyaltyCoupon.used_at.between(start, end),
            )
        ) or 0

        rewards_used = await self.db.scalar(
            select(func.count(func.distinct(LoyaltyCoupon.reward_id))).where(
                LoyaltyCoupon.company_id == company_id,
                LoyaltyCoupon.is_used == True,
                LoyaltyCoupon.used_at.between(start, end),
                LoyaltyCoupon.reward_id.is_not(None),
            )
        ) or 0

        avg_clv_xof = await self.db.scalar(
            select(func.coalesce(func.avg(CustomerScore.total_spent_xof), 0)).where(
                CustomerScore.company_id == company_id
            )
        ) or 0

        return {
            "period": period,
            "date_from": start.date().isoformat(),
            "date_to": end.date().isoformat(),
            "orders_count": int(orders_count),
            "orders_delivered": int(orders_delivered),
            "orders_cancelled": int(orders_cancelled),
            "revenue_xof": int(revenue_result),
            "avg_order_xof": round(avg_order),
            "unique_customers": int(unique_customers),
            "new_customers": int(new_customers),
            "avg_customer_value_xof": round(revenue_result / unique_customers) if unique_customers else 0,
            "avg_clv_xof": round(float(avg_clv_xof)),
            "points_distributed": int(points_distributed),
            "coupons_used": int(coupons_used),
            "rewards_used": int(rewards_used),
            "customer_segments": await self._get_segment_counts(company_id),
            "top_products": top_products,
            "payment_by_operator": payment_by_operator,
            "top_customers": await self._get_top_customers(company_id, start, end),
        }

    async def _get_orders_data(self, company_id: UUID, start: datetime, end: datetime) -> list[dict]:
        result = await self.db.execute(
            select(
                Order.order_number,
                Order.status,
                Order.type,
                Order.total_xof,
                Order.created_at,
                Payment.operator,
                Payment.transaction_ref,
                Payment.status.label("payment_status"),
                Payment.confirmed_at,
            )
            .outerjoin(Payment, Payment.order_id == Order.id)
            .where(
                Order.company_id == company_id,
                Order.created_at.between(start, end),
            )
            .order_by(Order.created_at.desc())
        )
        rows = result.all()
        return [
            {
                "order_number": row.order_number,
                "status": row.status,
                "type": row.type,
                "total_xof": row.total_xof,
                "created_at": row.created_at.strftime("%d/%m/%Y %H:%M") if row.created_at else "",
                "payment_operator": row.operator or "",
                "payment_ref": row.transaction_ref or "",
                "payment_status": row.payment_status or "",
                "confirmed_at": row.confirmed_at.strftime("%d/%m/%Y %H:%M") if row.confirmed_at else "",
            }
            for row in rows
        ]

    async def export_csv(
        self,
        company_id: UUID,
        period: str = "month",
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
    ) -> bytes:
        if date_from and date_to:
            start = datetime.combine(date_from, datetime.min.time()).replace(tzinfo=timezone.utc)
            end = datetime.combine(date_to, datetime.max.time()).replace(tzinfo=timezone.utc)
        else:
            start, end = self._period_bounds(period)

        rows = await self._get_orders_data(company_id, start, end)

        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=[
                "order_number",
                "status",
                "type",
                "total_xof",
                "created_at",
                "payment_operator",
                "payment_ref",
                "payment_status",
                "confirmed_at",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)
        return output.getvalue().encode("utf-8-sig")

    async def export_excel(
        self,
        company_id: UUID,
        period: str = "month",
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
    ) -> bytes:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        if date_from and date_to:
            start = datetime.combine(date_from, datetime.min.time()).replace(tzinfo=timezone.utc)
            end = datetime.combine(date_to, datetime.max.time()).replace(tzinfo=timezone.utc)
        else:
            start, end = self._period_bounds(period)

        rows = await self._get_orders_data(company_id, start, end)
        summary = await self.get_summary(company_id, period, date_from, date_to)

        wb = openpyxl.Workbook()

        ws_summary = wb.active
        ws_summary.title = "Resume"
        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(color="FFFFFF", bold=True)

        summary_data = [
            ("Periode", f"{summary['date_from']} -> {summary['date_to']}"),
            ("Commandes totales", summary["orders_count"]),
            ("Commandes livrees", summary["orders_delivered"]),
            ("Commandes annulees", summary["orders_cancelled"]),
            ("Chiffre d'affaires (FCFA)", summary["revenue_xof"]),
            ("Panier moyen (FCFA)", summary["avg_order_xof"]),
            ("Clients uniques", summary["unique_customers"]),
            ("Nouveaux clients", summary["new_customers"]),
            ("Points distribues", summary["points_distributed"]),
            ("Coupons utilises", summary["coupons_used"]),
        ]
        for index, (label, value) in enumerate(summary_data, start=1):
            ws_summary.cell(row=index, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=index, column=2, value=value)
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 25

        ws_orders = wb.create_sheet("Commandes")
        headers = [
            "No Commande",
            "Statut",
            "Type",
            "Total (FCFA)",
            "Creee le",
            "Operateur",
            "Reference paiement",
            "Statut paiement",
            "Confirmee le",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws_orders.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_index, row in enumerate(rows, start=2):
            ws_orders.cell(row=row_index, column=1, value=row["order_number"])
            ws_orders.cell(row=row_index, column=2, value=row["status"])
            ws_orders.cell(row=row_index, column=3, value=row["type"])
            ws_orders.cell(row=row_index, column=4, value=row["total_xof"])
            ws_orders.cell(row=row_index, column=5, value=row["created_at"])
            ws_orders.cell(row=row_index, column=6, value=row["payment_operator"])
            ws_orders.cell(row=row_index, column=7, value=row["payment_ref"])
            ws_orders.cell(row=row_index, column=8, value=row["payment_status"])
            ws_orders.cell(row=row_index, column=9, value=row["confirmed_at"])

        for col in ws_orders.columns:
            ws_orders.column_dimensions[col[0].column_letter].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def export_pdf(
        self,
        company_id: UUID,
        company_name: str,
        period: str = "month",
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
    ) -> bytes:
        import html as html_module

        if date_from and date_to:
            start = datetime.combine(date_from, datetime.min.time()).replace(tzinfo=timezone.utc)
            end = datetime.combine(date_to, datetime.max.time()).replace(tzinfo=timezone.utc)
        else:
            start, end = self._period_bounds(period)

        summary = await self.get_summary(company_id, period, date_from, date_to)
        rows = await self._get_orders_data(company_id, start, end)

        escape = html_module.escape
        rows_html = "".join(
            f"<tr>"
            f"<td>{escape(row['order_number'])}</td>"
            f"<td>{escape(row['status'])}</td>"
            f"<td>{escape(row['type'])}</td>"
            f"<td class='right'>{row['total_xof']:,} FCFA</td>"
            f"<td>{escape(row['created_at'])}</td>"
            f"<td>{escape(row['payment_operator'])}</td>"
            f"<td>{escape(row['payment_status'])}</td>"
            f"</tr>"
            for row in rows
        )

        html = f"""<!DOCTYPE html>
<html lang="fr">
<head><meta charset="UTF-8">
<style>
  body {{font-family: Arial, sans-serif; font-size: 12px; color: #1a1a1a;}}
  h1 {{color: #2563eb; font-size: 18px;}}
  h2 {{color: #374151; font-size: 14px; border-bottom: 1px solid #e5e7eb; padding-bottom: 4px;}}
  table {{width: 100%; border-collapse: collapse; margin-top: 8px;}}
  th {{background: #2563eb; color: white; padding: 6px 8px; text-align: left; font-size: 11px;}}
  td {{padding: 5px 8px; border-bottom: 1px solid #f3f4f6; font-size: 11px;}}
  .right {{text-align: right;}}
  .kpi {{display: inline-block; margin: 8px 16px 8px 0; padding: 8px 16px; background: #f3f4f6; border-radius: 8px;}}
  .kpi-val {{font-size: 20px; font-weight: bold; color: #2563eb;}}
  .kpi-lbl {{font-size: 11px; color: #6b7280;}}
</style>
</head>
<body>
<h1>Rapport - {escape(company_name)}</h1>
<p>Periode : {summary['date_from']} -> {summary['date_to']}</p>

<h2>Indicateurs cles</h2>
<div class="kpi"><div class="kpi-val">{summary['orders_count']}</div><div class="kpi-lbl">Commandes</div></div>
<div class="kpi"><div class="kpi-val">{summary['orders_delivered']}</div><div class="kpi-lbl">Livrees</div></div>
<div class="kpi"><div class="kpi-val">{summary['revenue_xof']:,} FCFA</div><div class="kpi-lbl">CA confirme</div></div>
<div class="kpi"><div class="kpi-val">{summary['avg_order_xof']:,} FCFA</div><div class="kpi-lbl">Panier moyen</div></div>
<div class="kpi"><div class="kpi-val">{summary['unique_customers']}</div><div class="kpi-lbl">Clients</div></div>

<h2>Commandes ({len(rows)})</h2>
<table>
  <thead>
    <tr>
      <th>No Commande</th><th>Statut</th><th>Type</th>
      <th>Total</th><th>Date</th><th>Operateur</th><th>Paiement</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
</table>

<p style="margin-top:24px;font-size:10px;color:#9ca3af;">
  Genere le {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC - SmartCheckout / Fiissa
</p>
</body>
</html>"""

        try:
            from weasyprint import HTML

            return HTML(string=html).write_pdf()
        except Exception:
            return html.encode("utf-8")

    async def generate_monthly_report(self, company_id: UUID, year: int, month: int) -> dict:
        _, last_day = monthrange(year, month)
        start = datetime(year, month, 1, tzinfo=timezone.utc)
        end = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)

        rows = await self._get_orders_data(company_id, start, end)
        orders_confirmed = sum(1 for row in rows if row["payment_status"] == "confirmed")
        revenue = sum(row["total_xof"] for row in rows if row["payment_status"] == "confirmed")

        return {
            "company_id": str(company_id),
            "year": year,
            "month": month,
            "orders_total": len(rows),
            "orders_confirmed": orders_confirmed,
            "revenue_xof": revenue,
        }
